#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일 생성 모듈
파일명: utils/excel_handler.py
작성자: 청약 자동화 시스템
설명: 분양정보 데이터를 엑셀 파일로 생성 및 스타일링
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
from typing import List, Dict, Optional

class ExcelHandler:
    """엑셀 파일 생성 및 스타일링 클래스"""
    
    def __init__(self, output_dir: str = "output"):
        """
        초기화
        Args:
            output_dir (str): 출력 디렉토리
        """
        self.output_dir = output_dir
        self.logger = logging.getLogger(__name__)
        
        # 출력 디렉토리 생성
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.logger.info(f"출력 디렉토리 생성: {output_dir}")
        
        # 엑셀 스타일 정의
        self._setup_styles()
    
    def _setup_styles(self):
        """엑셀 스타일 설정"""
        # 헤더 스타일
        self.header_font = Font(
            name='맑은 고딕',
            size=11,
            bold=True,
            color='FFFFFF'
        )
        
        self.header_fill = PatternFill(
            start_color='4F81BD',
            end_color='4F81BD',
            fill_type='solid'
        )
        
        # 데이터 스타일
        self.data_font = Font(
            name='맑은 고딕',
            size=10
        )
        
        # 테두리 스타일
        thin_border = Side(border_style="thin", color="000000")
        self.border = Border(
            top=thin_border,
            left=thin_border,
            right=thin_border,
            bottom=thin_border
        )
        
        # 정렬 스타일
        self.center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        
        self.left_alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )
    
    def create_excel_file(self, data: Dict[str, List[Dict]], filename: Optional[str] = None) -> str:
        """
        분양정보 엑셀 파일 생성
        Args:
            data (Dict[str, List[Dict]]): 분양정보 데이터
            filename (Optional[str]): 파일명 (None이면 자동 생성)
        Returns:
            str: 생성된 파일의 전체 경로
        """
        try:
            # 파일명 생성
            if filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"청약분양정보_{timestamp}.xlsx"
            
            filepath = os.path.join(self.output_dir, filename)
            
            self.logger.info(f"엑셀 파일 생성 시작: {filename}")
            
            # 워크북 생성
            wb = Workbook()
            
            # 기본 시트 제거
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 데이터가 있는 경우에만 시트 생성
            total_count = 0
            
            # 일반 분양정보 시트
            if data.get('general') and len(data['general']) > 0:
                self._create_worksheet(wb, '일반분양', data['general'])
                total_count += len(data['general'])
            
            # APT 분양정보 시트
            if data.get('apt') and len(data['apt']) > 0:
                self._create_worksheet(wb, 'APT분양', data['apt'])
                total_count += len(data['apt'])
            
            # 오피스텔 분양정보 시트
            if data.get('officetel') and len(data['officetel']) > 0:
                self._create_worksheet(wb, '오피스텔분양', data['officetel'])
                total_count += len(data['officetel'])
            
            # 통합 시트 생성 (모든 데이터)
            all_data = []
            for category, items in data.items():
                for item in items:
                    # 카테고리 정보 추가
                    item_with_category = item.copy()
                    item_with_category['분양유형'] = {
                        'general': '일반분양',
                        'apt': 'APT분양',
                        'officetel': '오피스텔분양'
                    }.get(category, category)
                    all_data.append(item_with_category)
            
            if all_data:
                self._create_worksheet(wb, '전체분양정보', all_data)
            
            # 요약 시트 생성
            self._create_summary_worksheet(wb, data)
            
            # 파일 저장
            wb.save(filepath)
            
            self.logger.info(f"엑셀 파일 생성 완료: {filepath} (총 {total_count}건)")
            return filepath
            
        except Exception as e:
            self.logger.error(f"엑셀 파일 생성 오류: {e}")
            raise
    
    def _create_worksheet(self, workbook: Workbook, sheet_name: str, data: List[Dict]):
        """
        워크시트 생성 및 데이터 입력
        Args:
            workbook (Workbook): 워크북 객체
            sheet_name (str): 시트명
            data (List[Dict]): 입력할 데이터
        """
        try:
            # 워크시트 생성
            ws = workbook.create_sheet(title=sheet_name)
            
            if not data:
                # 데이터가 없는 경우
                ws['A1'] = "데이터가 없습니다."
                ws['A1'].font = self.data_font
                return
            
            # 데이터프레임 생성
            df = pd.DataFrame(data)
            
            # 순번 컬럼 추가 (맨 앞에)
            df.insert(0, '순번', range(1, len(df) + 1))
            
            # 컬럼 순서 정리
            desired_columns = [
                '순번', '분양유형', '주택명', '공급지역', '공급규모', '모집공고일',
                '청약접수시작일', '청약접수종료일', '당첨발표일', '계약시작일', 
                '계약종료일', '입주예정일', '분양가격', '건설업체', '분양업체',
                '연락처', '홈페이지', '주택형별정보', '특별공급', '일반공급'
            ]
            
            # 존재하는 컬럼만 선택
            available_columns = [col for col in desired_columns if col in df.columns]
            df = df[available_columns]
            
            # 데이터를 워크시트에 입력
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 스타일 적용
                    if r_idx == 1:  # 헤더 행
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.center_alignment
                    else:  # 데이터 행
                        cell.font = self.data_font
                        # 숫자 컬럼은 중앙 정렬, 텍스트는 좌측 정렬
                        if c_idx == 1 or '일' in str(df.columns[c_idx-1]):  # 순번, 날짜 컬럼
                            cell.alignment = self.center_alignment
                        else:
                            cell.alignment = self.left_alignment
                    
                    cell.border = self.border
            
            # 컬럼 너비 자동 조정
            self._adjust_column_width(ws)
            
            # 필터 적용
            if len(df) > 0:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"
            
            # 틀 고정 (헤더 행)
            ws.freeze_panes = ws['A2']
            
            self.logger.info(f"워크시트 '{sheet_name}' 생성 완료: {len(data)}건")
            
        except Exception as e:
            self.logger.error(f"워크시트 생성 오류 ({sheet_name}): {e}")
            raise
    
    def _create_summary_worksheet(self, workbook: Workbook, data: Dict[str, List[Dict]]):
        """
        요약 워크시트 생성
        Args:
            workbook (Workbook): 워크북 객체
            data (Dict[str, List[Dict]]): 분양정보 데이터
        """
        try:
            ws = workbook.create_sheet(title='요약정보', index=0)  # 첫 번째 시트로 설정
            
            # 제목
            ws['A1'] = '청약 분양정보 요약'
            ws['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
            ws.merge_cells('A1:D1')
            
            # 생성 일시
            ws['A3'] = '생성 일시:'
            ws['B3'] = datetime.now().strftime('%Y년 %m월 %d일 %H시 %M분')
            
            # 카테고리별 통계
            ws['A5'] = '분양 유형별 현황'
            ws['A5'].font = Font(name='맑은 고딕', size=12, bold=True)
            
            # 통계 테이블 헤더
            headers = ['분양유형', '건수', '비율(%)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            # 통계 데이터
            categories = {
                '일반분양': len(data.get('general', [])),
                'APT분양': len(data.get('apt', [])),
                '오피스텔분양': len(data.get('officetel', []))
            }
            
            total_count = sum(categories.values())
            
            row = 7
            for category, count in categories.items():
                if count > 0:  # 데이터가 있는 경우만 표시
                    percentage = (count / total_count * 100) if total_count > 0 else 0
                    
                    ws.cell(row=row, column=1, value=category).border = self.border
                    ws.cell(row=row, column=2, value=count).border = self.border
                    ws.cell(row=row, column=3, value=f"{percentage:.1f}").border = self.border
                    
                    # 중앙 정렬
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).alignment = self.center_alignment
                        ws.cell(row=row, column=col).font = self.data_font
                    
                    row += 1
            
            # 총계
            ws.cell(row=row, column=1, value='총계').border = self.border
            ws.cell(row=row, column=2, value=total_count).border = self.border
            ws.cell(row=row, column=3, value='100.0').border = self.border
            
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.alignment = self.center_alignment
                cell.font = Font(name='맑은 고딕', size=10, bold=True)
                cell.border = self.border
            
            # 최근 분양정보 (상위 10개)
            ws[f'A{row+3}'] = '최근 분양정보 (상위 10개)'
            ws[f'A{row+3}'].font = Font(name='맑은 고딕', size=12, bold=True)
            
            # 모든 데이터 수집 및 정렬
            all_items = []
            for category, items in data.items():
                for item in items:
                    item_copy = item.copy()
                    item_copy['분양유형'] = {
                        'general': '일반분양',
                        'apt': 'APT분양', 
                        'officetel': '오피스텔분양'
                    }.get(category, category)
                    all_items.append(item_copy)
            
            # 모집공고일 기준으로 정렬 (최신순)
            try:
                all_items.sort(key=lambda x: x.get('모집공고일', ''), reverse=True)
            except:
                pass  # 정렬 실패 시 원본 순서 유지
            
            # 상위 10개 표시
            recent_headers = ['순번', '분양유형', '주택명', '공급지역', '모집공고일', '청약접수시작일']
            header_row = row + 4
            
            for col, header in enumerate(recent_headers, 1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.center_alignment
                cell.border = self.border
            
            for i, item in enumerate(all_items[:10], 1):
                data_row = header_row + i
                values = [
                    i,
                    item.get('분양유형', ''),
                    item.get('주택명', ''),
                    item.get('공급지역', ''),
                    item.get('모집공고일', ''),
                    item.get('청약접수시작일', '')
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=data_row, column=col, value=value)
                    cell.font = self.data_font
                    cell.alignment = self.center_alignment if col == 1 else self.left_alignment
                    cell.border = self.border
            
            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 25
            
            self.logger.info("요약 워크시트 생성 완료")
            
        except Exception as e:
            self.logger.error(f"요약 워크시트 생성 오류: {e}")
            # 요약 시트 생성 실패는 전체 작업을 중단시키지 않음
    
    def _adjust_column_width(self, worksheet):
        """
        컬럼 너비 자동 조정
        Args:
            worksheet: 워크시트 객체
        """
        try:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                # 최소 10, 최대 50으로 제한
                adjusted_width = min(max(length + 2, 10), 50)
                worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width
                
        except Exception as e:
            self.logger.warning(f"컬럼 너비 조정 실패: {e}")
    
    def create_test_excel(self) -> str:
        """
        테스트용 엑셀 파일 생성
        Returns:
            str: 생성된 테스트 파일 경로
        """
        test_data = {
            'general': [
                {
                    '주택명': '테스트 아파트 1단지',
                    '공급지역': '서울시 강남구',
                    '공급규모': '500세대',
                    '모집공고일': '2024-01-15',
                    '청약접수시작일': '2024-01-20',
                    '청약접수종료일': '2024-01-22',
                    '당첨발표일': '2024-01-25',
                    '계약시작일': '2024-02-01',
                    '입주예정일': '2025-12-01',
                    '분양가격': '5억원~8억원',
                    '홈페이지': 'http://test-apt.co.kr',
                    '연락처': '02-1234-5678',
                    '건설업체': '테스트건설',
                    '분양업체': '테스트부동산'
                }
            ],
            'apt': [],
            'officetel': []
        }
        
        return self.create_excel_file(test_data, "테스트_청약분양정보.xlsx")